"""
Configuration classes for QA Report Helper package.
"""

from dataclasses import dataclass
from typing import Tuple
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


@dataclass(frozen=True)
class Config:
    """Configuration constants for the QA Report application."""
    REQUIRED_COLUMNS: Tuple[str, ...] = ("Lead Status", "Agent Name", "DQ Reason")
    OPTIONAL_COLUMNS: Tuple[str, ...] = ("Segment Tagging", "JT Persona Tagging")
    ACCEPTED_LEAD_STATUS: Tuple[str, ...] = ("Qualified", "Disqualified")
    ACCEPTED_DQ_REASON: Tuple[str, ...] = (
        "Not An RPC", "Invalid Phone Number", "Invalid Job Title", "Invalid Details", "Email Bounce Back", 
        "Gmail Bounce Back", "Invalid Email", "Toll Free Number", "Invalid Geo", 
        "Invalid Industry", "Invalid Employee Size", "Invalid Revenue Size", 
        "Same Prospect Duplicate", "Client Suppression", "Suspect Profile",
        "Internal Suppression", "Dead Contact", "Not In TAL", "Extra CPC", 
        "Invalid Zipcode", "Invalid Address"
    )
    MAX_FILE_SIZE_MB: int = 50
    SUPPORTED_EXTENSIONS: Tuple[str, ...] = ('xlsx', 'xlsm')


@dataclass(frozen=True)
class ExcelStyling:
    """Excel styling configuration."""
    header_fill: PatternFill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font: Font = Font(bold=True)
    grand_total_fill: PatternFill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    grand_total_font: Font = Font(bold=True)
    center_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    thin_border: Border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )