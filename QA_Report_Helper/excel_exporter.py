"""
Excel export functionality for QA Report Helper package.
Updated to include date information in reports.
"""

import io
from datetime import datetime, date
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .config import ExcelStyling


class ExcelExporter:
    """Handles Excel file generation with professional formatting."""
    
    def __init__(self, styling: ExcelStyling):
        self.styling = styling
    
    def create_excel_report(
        self,
        reports: Dict[str, List[List[Any]]],
        campaign_id: str = "",
        selected_date: date = None,
        multi_level_defined_combos: List[tuple] = None
    ) -> bytes:
        """Create formatted Excel report with email content and date information.

        multi_level_defined_combos: the combinations defined in the Mapping
        sheet. When provided, rows in the Custom Multi-Level Report whose
        combination is NOT among them (out-of-snapshot pairings) are highlighted.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "QA_Report"
        
        current_row = 1
        
        # Add email content at the top if campaign_id is provided
        if campaign_id:
            date_str = selected_date.strftime("%d-%b-%y") if selected_date else self._format_today_date()
            
            # Email greeting and header
            ws.cell(row=current_row, column=1, value="Hi Team,").font = Font(bold=False, size=11)
            current_row += 2
            
            ws.cell(row=current_row, column=1, value=f"PFB QA_Report_{campaign_id}_{date_str}").font = Font(bold=True, size=12)
            current_row += 2
        
        # Add Combined QA Report first
        if "Combined QA Report" in reports:
            current_row = self._add_report_to_worksheet(ws, reports["Combined QA Report"], current_row)
            current_row += 1
        
        # Add optional reports next (Segment and JT Persona)
        optional_reports = ["Segment Wise Qualified Count", "JT Persona Wise Qualified Count"]
        for report_name in optional_reports:
            if report_name in reports:
                current_row = self._add_report_to_worksheet(ws, reports[report_name], current_row)
                current_row += 1  # Add spacing between reports
        
        # Add Custom Multi-Level Report with hierarchical Segment merging
        # for every grouping column except the innermost (which is unique
        # per row within its group)
        if "Custom Multi-Level Report" in reports:
            report_data = reports["Custom Multi-Level Report"]
            if report_data and len(report_data) > 1 and len(report_data[0]) >= 2:
                # Total columns = grouping columns + 1 (Count)
                num_grouping = len(report_data[0]) - 1
                # Merge all grouping columns except the innermost
                # e.g. 3 grouping cols -> merge cols [1, 2]; 2 -> [1]; 1 -> []
                merge_cols = list(range(1, num_grouping))
                defined_set = (
                    {tuple(c) for c in multi_level_defined_combos}
                    if multi_level_defined_combos is not None else None
                )
                current_row = self._add_grouped_report_to_worksheet(
                    ws, report_data, current_row, group_columns=merge_cols,
                    defined_combos=defined_set
                )
            else:
                # Fall back to plain rendering if shape is unexpected
                current_row = self._add_report_to_worksheet(
                    ws, report_data, current_row
                )
            current_row += 1
        
        # Add core reports last (Agent Wise Summary, Primary Reason Disqualified)
        core_reports = ["Agent Wise Summary", "Primary Reason Disqualified"]
        for report_name in core_reports:
            if report_name in reports:
                current_row = self._add_report_to_worksheet(ws, reports[report_name], current_row)
                current_row += 1  # Add spacing between reports
        
        # Add summary at the end if campaign_id is provided
        if campaign_id:
            ws.cell(row=current_row, column=1, value="Best regards,").font = Font(bold=False, size=11)
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _format_today_date(self) -> str:
        """Format today's date in the required format (e.g., 2nd-Aug-2025)."""
        today = datetime.now()
        
        # Get day with ordinal suffix
        day = today.day
        if 10 <= day % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        
        # Format: 2nd-Aug-2025
        formatted_date = f"{day}{suffix}-{today.strftime('%b')}-{today.year}"
        return formatted_date
    
    def _add_report_to_worksheet(self, ws: Worksheet, report_data: List[List[Any]], start_row: int) -> int:
        """Add a single report to the worksheet with formatting."""
        current_row = start_row
        
        for row_idx, row_data in enumerate(report_data):
            # Add data to cells
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                self._apply_cell_formatting(cell, row_idx, len(report_data))
            
            current_row += 1
        
        return current_row
    
    def _add_grouped_report_to_worksheet(
        self,
        ws: Worksheet,
        report_data: List[List[Any]],
        start_row: int,
        group_columns: List[int] = None,
        defined_combos: set = None
    ) -> int:
        """
        Add a report to the worksheet and merge consecutive equal cells in
        each of the specified 1-indexed grouping columns. Merging is
        HIERARCHICAL: a cell in group_columns[k] is merged across rows
        only when every column in group_columns[:k+1] is equal across
        those same rows.
        
        This produces a nested grouped look like:
        
            | Country | Segment | Persona     | Count |
            | US      | Seg A   | Persona 1   | 28    |
            |         |         | Persona 2   | 40    |  (Country + Segment merged)
            |         | Seg B   | Persona 1   | 12    |  (Country merged only)
            | Canada  | Seg A   | Persona 1   | 5     |
        
        Header row and Grand Total row (if present) are excluded from
        merging. Passing an empty group_columns list writes the report
        without any merging.
        """
        if group_columns is None:
            group_columns = [1]
        
        current_row = start_row
        total_rows = len(report_data)
        
        # Write all cells first (identical to _add_report_to_worksheet)
        for row_idx, row_data in enumerate(report_data):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                self._apply_cell_formatting(cell, row_idx, total_rows)
            current_row += 1

        # Highlight rows whose combination is NOT defined in the Mapping
        # snapshot (out-of-snapshot pairings). Applied here — before any
        # merging — so it works whether or not columns are merged. Only the
        # innermost grouping cell + the count cell are highlighted, since the
        # outer grouping cells get merged (and would hide a per-row fill).
        if defined_combos is not None and report_data:
            num_grouping = len(report_data[0]) - 1
            highlight_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            highlight_font = Font(color="9C0006", bold=True)
            for row_idx in range(1, total_rows):  # skip header
                row_vals = report_data[row_idx]
                if not row_vals or str(row_vals[0]).strip().lower() == "grand total":
                    continue
                combo = tuple(row_vals[:num_grouping])
                if combo in defined_combos:
                    continue
                ws_row = start_row + row_idx
                for col in (num_grouping, num_grouping + 1):
                    cell = ws.cell(row=ws_row, column=col)
                    cell.fill = highlight_fill
                    cell.font = highlight_font

        if not group_columns:
            return current_row  # Nothing to merge
        
        # Merge boundaries
        first_data_row = start_row + 1  # skip header
        last_written_row = current_row - 1
        
        # Detect Grand Total in the last row (first-column value)
        last_row_first_cell = ws.cell(row=last_written_row, column=1).value
        has_grand_total = (
            isinstance(last_row_first_cell, str)
            and last_row_first_cell.strip().lower() == "grand total"
        )
        last_data_row = last_written_row - 1 if has_grand_total else last_written_row
        
        if last_data_row < first_data_row:
            return current_row  # No data rows to merge
        
        # For each grouping column, merge consecutive rows where ALL
        # columns in group_columns up to and including this one are equal.
        for merge_col in group_columns:
            # Prefix = every grouping column at or before this one
            prefix_cols = [c for c in group_columns if c <= merge_col]

            def prefix_key(row_num, cols=prefix_cols):
                # Read from the ORIGINAL report_data, not the worksheet: merging
                # an outer column blanks that column's cells (openpyxl keeps the
                # value only in the merged top-left), which would corrupt this
                # comparison for inner columns and split runs incorrectly.
                row_vals = report_data[row_num - start_row]
                return tuple(row_vals[c - 1] for c in cols)
            
            merge_start = first_data_row
            while merge_start <= last_data_row:
                base_key = prefix_key(merge_start)
                merge_end = merge_start
                while (
                    merge_end + 1 <= last_data_row
                    and prefix_key(merge_end + 1) == base_key
                ):
                    merge_end += 1
                
                if merge_end > merge_start:
                    ws.merge_cells(
                        start_row=merge_start,
                        start_column=merge_col,
                        end_row=merge_end,
                        end_column=merge_col
                    )
                    # Re-apply center alignment + border to merged top-left
                    top_cell = ws.cell(row=merge_start, column=merge_col)
                    top_cell.alignment = self.styling.center_alignment
                    top_cell.border = self.styling.thin_border
                
                merge_start = merge_end + 1
        
        return current_row
    
    def _apply_cell_formatting(self, cell, row_idx: int, total_rows: int) -> None:
        """Apply formatting to a cell based on its position."""
        cell.alignment = self.styling.center_alignment
        cell.border = self.styling.thin_border
        
        # Header row formatting
        if row_idx == 0:
            cell.fill = self.styling.header_fill
            cell.font = self.styling.header_font
        
        # Grand Total row formatting
        elif (isinstance(cell.value, str) and cell.value.lower().strip() == "grand total") or \
             (row_idx == total_rows - 1 and total_rows > 2):
            cell.fill = self.styling.grand_total_fill
            cell.font = self.styling.grand_total_font
    
    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set width with padding
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width